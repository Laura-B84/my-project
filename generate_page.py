"""Генерирует docs/index.html — статическую страницу-обзор со сводкой
отзывов из otzyvy_halyk.xlsx, для публикации через GitHub Pages.

Повторяет структуру xlsx: блоки по дням, Android слева, Apple справа.
"""

import html
from datetime import datetime
from itertools import groupby
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
XLSX_FILE = BASE_DIR / "otzyvy_halyk.xlsx"
OUT_FILE = BASE_DIR / "docs" / "index.html"

FIELDS = [
    "Дата отзыва", "Страна/сторфронт", "Рейтинг",
    "Заголовок отзыва", "Текст отзыва", "Автор", "Версия приложения",
    "ID/хэш отзыва", "Негативный",
]
DATE_IDX = FIELDS.index("Дата отзыва")
ID_IDX = FIELDS.index("ID/хэш отзыва")
NEG_IDX = FIELDS.index("Негативный")
LEFT_COL = 1
RIGHT_COL = LEFT_COL + len(FIELDS) + 1

wb = load_workbook(XLSX_FILE)
ws = wb.active

android_rows, apple_rows = [], []
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row:
        continue
    left = list(row[LEFT_COL - 1:LEFT_COL - 1 + len(FIELDS)])
    right = list(row[RIGHT_COL - 1:RIGHT_COL - 1 + len(FIELDS)])
    if left[ID_IDX]:
        android_rows.append(left)
    if right[ID_IDX]:
        apple_rows.append(right)

total = len(android_rows) + len(apple_rows)
negative = sum(1 for r in android_rows + apple_rows if r[NEG_IDX] == "да")


def esc(v):
    return html.escape(str(v)) if v is not None else ""


def review_row_html(r):
    date, country, rating, title, text, author, version, rid, is_neg = r
    cls = ' class="neg"' if is_neg == "да" else ""
    date_str = date.strftime("%Y-%m-%d %H:%M") if hasattr(date, "strftime") else esc(date)
    return (
        f"<tr{cls}><td>{date_str}</td><td>{esc(country)}</td>"
        f"<td>{esc(rating)}★</td><td>{esc(title)}</td><td>{esc(text)}</td>"
        f"<td>{esc(author)}</td><td>{esc(version)}</td></tr>"
    )


def platform_table_html(name, rows):
    if not rows:
        return f'<div class="platform-col empty"><h3>{esc(name)}</h3><p class="none">нет отзывов за день</p></div>'
    table_rows = "".join(review_row_html(r) for r in sorted(rows, key=lambda r: r[DATE_IDX] or datetime.min, reverse=True))
    return f"""
      <div class="platform-col">
        <h3>{esc(name)} <span class="count">{len(rows)}</span></h3>
        <div class="wrap">
          <table>
            <thead><tr><th>Дата</th><th>Страна</th><th>Рейтинг</th><th>Заголовок</th><th>Текст</th><th>Автор</th><th>Версия</th></tr></thead>
            <tbody>{table_rows}</tbody>
          </table>
        </div>
      </div>"""


def group_by_day(rows):
    rows = sorted(rows, key=lambda r: r[DATE_IDX] or datetime.min, reverse=True)
    by_day = {}
    for day, day_rows in groupby(rows, key=lambda r: r[DATE_IDX].date() if hasattr(r[DATE_IDX], "date") else None):
        by_day[day] = list(day_rows)
    return by_day


android_by_day = group_by_day(android_rows)
apple_by_day = group_by_day(apple_rows)
all_days = sorted(set(android_by_day) | set(apple_by_day), key=lambda d: d or datetime.min.date(), reverse=True)

blocks_html = []
for day in all_days:
    day_label = day.strftime("%d.%m.%Y") if day else "Без даты"
    blocks_html.append(f"""
      <section class="day-block">
        <h2>{day_label}</h2>
        <div class="side-by-side">
          {platform_table_html("Google Play (Android)", android_by_day.get(day, []))}
          {platform_table_html("App Store (iOS)", apple_by_day.get(day, []))}
        </div>
      </section>""")

html_out = f"""<title>Halyk Kazakhstan — Отзывы</title>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
  :root {{
    --bg: #ffffff; --fg: #1a1a1a; --muted: #666; --border: #e2e2e2;
    --neg-bg: #fdeaea; --accent: #0a7cff; --block-bg: #f7f9fc; --store-bg: #f0f0f0;
  }}
  @media (prefers-color-scheme: dark) {{
    :root {{ --bg: #14161a; --fg: #eaeaea; --muted: #9a9a9a; --border: #2a2d33; --neg-bg: #3a1f22; --accent: #5eaaff; --block-bg: #1b1e24; --store-bg: #20242b; }}
  }}
  * {{ box-sizing: border-box; }}
  body {{ background: var(--bg); color: var(--fg); font-family: -apple-system, Segoe UI, Roboto, sans-serif; margin: 0; padding: 2rem 1.25rem 4rem; }}
  main {{ max-width: 1200px; margin: 0 auto; }}
  h1 {{ font-size: 1.6rem; margin-bottom: 0.25rem; }}
  .sub {{ color: var(--muted); margin-bottom: 1.5rem; }}
  .stats {{ display: flex; gap: 1rem; margin-bottom: 2rem; flex-wrap: wrap; }}
  .stat {{ border: 1px solid var(--border); border-radius: 10px; padding: 0.9rem 1.2rem; min-width: 140px; }}
  .stat b {{ display: block; font-size: 1.6rem; }}
  .stat span {{ color: var(--muted); font-size: 0.85rem; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.82rem; }}
  th, td {{ text-align: left; padding: 0.45rem 0.55rem; border-bottom: 1px solid var(--border); vertical-align: top; }}
  th {{ color: var(--muted); font-weight: 600; }}
  tr.neg {{ background: var(--neg-bg); }}
  .wrap {{ overflow-x: auto; border: 1px solid var(--border); border-radius: 10px; }}
  a {{ color: var(--accent); }}
  .day-block {{ background: var(--block-bg); border: 1px solid var(--border); border-radius: 12px; padding: 1rem 1.2rem 1.4rem; margin-bottom: 1.2rem; }}
  .day-block h2 {{ margin: 0 0 0.8rem; font-size: 1.15rem; }}
  .side-by-side {{ display: grid; grid-template-columns: 1fr 1fr; gap: 1rem; }}
  @media (max-width: 800px) {{ .side-by-side {{ grid-template-columns: 1fr; }} }}
  .platform-col h3 {{ background: var(--store-bg); border-radius: 6px; padding: 0.35rem 0.6rem; font-size: 0.9rem; margin: 0 0 0.5rem; display: inline-flex; gap: 0.4rem; align-items: center; }}
  .count {{ color: var(--muted); font-weight: 400; }}
  .platform-col.empty .none {{ color: var(--muted); font-size: 0.85rem; font-style: italic; }}
</style>
<main>
  <h1>Halyk Kazakhstan — Отзывы</h1>
  <div class="sub">Автосбор из App Store и Google Play (сторфронт KZ) · <a href="https://github.com/Laura-B84/my-project/blob/master/prd.md">PRD проекта</a></div>
  <div class="stats">
    <div class="stat"><b>{total}</b><span>всего отзывов</span></div>
    <div class="stat"><b>{negative}</b><span>негативных (≤3★)</span></div>
  </div>
  {''.join(blocks_html)}
</main>
"""

OUT_FILE.parent.mkdir(exist_ok=True)
OUT_FILE.write_text(html_out, encoding="utf-8")
print(f"Written {OUT_FILE} ({total} rows, {negative} negative)")
