"""
Сбор отзывов о приложении Halyk Kazakhstan (App Store + Google Play).

Логика соответствует prd.md (раздел 5): полностью локальный скрипт,
запускается ежедневно через Windows Task Scheduler, без облака.

- App Store: публичный RSS-фид (без ключей/авторизации).
- Google Play: публичный парсинг через google-play-scraper (без ключей).
- Дедупликация: App Store — по ID отзыва; Google Play — по хэшу
  автор+дата+текст (публичная страница не отдаёт стабильный ID).
- Ничего не теряет: новые отзывы добавляются к уже сохранённым, но лист
  каждый раз пересобирается заново в порядке от новых отзывов к старым.
- При появлении новых отзывов с рейтингом ≤3★ показывает Windows toast.
"""

import hashlib
import logging
import os
import time
from datetime import datetime, timezone
from itertools import groupby
from pathlib import Path

import requests
from google_play_scraper import Sort, reviews as gp_reviews
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_FILE = BASE_DIR / "otzyvy_halyk.xlsx"
LOG_FILE = BASE_DIR / "collect_reviews.log"

APP_STORE_ID = "440635615"
GOOGLE_PLAY_ID = "kz.kkb.homebank"

FIELDS = [
    "Дата отзыва", "Страна/сторфронт", "Рейтинг",
    "Заголовок отзыва", "Текст отзыва", "Автор", "Версия приложения",
    "ID/хэш отзыва", "Негативный",
]
DATE_IDX = FIELDS.index("Дата отзыва")
ID_IDX = FIELDS.index("ID/хэш отзыва")
NEG_IDX = FIELDS.index("Негативный")

# Два блока колонок бок о бок: Android (Google Play) слева, Apple (App
# Store) справа, с колонкой-разделителем между ними. Строки внутри одного
# дня выровнены по индексу — где данных с одной стороны не хватает,
# остаются пустые ячейки.
LEFT_COL = 1
GAP_COL = LEFT_COL + len(FIELDS)
RIGHT_COL = LEFT_COL + len(FIELDS) + 1
TOTAL_COLS = RIGHT_COL + len(FIELDS) - 1

# Старый (плоский) формат листа — нужен только для миграции уже
# накопленных данных при первом запуске с новым кодом.
OLD_HEADERS = [
    "Дата отзыва", "Стор", "Страна/сторфронт", "Рейтинг",
    "Заголовок отзыва", "Текст отзыва", "Автор", "Версия приложения",
    "ID/хэш отзыва", "Негативный",
]

NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

COUNTRIES = ["kz"]

# Ограничение по просьбе пользователя: не тянуть всю историю, а только
# последние N отзывов по каждому источнику при первом запуске.
REVIEW_LIMIT = 30

# Обычные запуски теперь идут только по будням (см. workflow) — после
# выходных между пятницей и понедельником проходит до 3 дней. Берём
# отзывы по дате (не по фиксированному количеству), с запасом на случай
# длинных выходных/праздников, чтобы ничего не потерялось.
DAILY_LOOKBACK_DAYS = 4

# Хранить в файле только отзывы за последние N дней — более старые
# записи удаляются при каждом запуске.
RETENTION_DAYS = 30

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    encoding="utf-8",
)


def hash_review(author, date, text):
    raw = f"{author}|{date}|{text}".encode("utf-8")
    return hashlib.sha256(raw).hexdigest()[:24]


def _label(value, default=""):
    """RSS-поля Apple обычно {"label": "..."}, но иногда приходят голой строкой."""
    if isinstance(value, dict):
        return value.get("label", default)
    if value is None:
        return default
    return value


def fetch_app_store_reviews(country, seen_ids, limit=REVIEW_LIMIT, cutoff_date=None):
    """Без `cutoff_date`: берёт top-`limit` самых свежих отзывов (обычный
    ежедневный режим). С `cutoff_date`: игнорирует `limit` и листает все
    доступные страницы (Apple отдаёт максимум 10 × 50 = 500 отзывов),
    оставляя только отзывы не старше `cutoff_date` — режим для разовой
    догрузки истории на глубину N дней."""
    raw_entries = []
    for page in range(1, 11):
        url = (
            f"https://itunes.apple.com/{country}/rss/customerreviews/"
            f"id={APP_STORE_ID}/sortBy=mostRecent/page={page}/json"
        )
        try:
            resp = requests.get(url, timeout=15)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            logging.warning("App Store %s page %s: %s", country, page, e)
            break

        entries = data.get("feed", {}).get("entry", [])
        reviews_on_page = [e for e in entries if "im:rating" in e]
        if not reviews_on_page:
            break
        raw_entries.extend(reviews_on_page)
        if cutoff_date is None and len(raw_entries) >= limit:
            break
        time.sleep(0.1)

    candidates = raw_entries if cutoff_date is not None else raw_entries[:limit]

    results = []
    for e in candidates:
        try:
            rid = _label(e.get("id"))
            if not rid or rid in seen_ids:
                continue
            try:
                dt = datetime.fromisoformat(_label(e.get("updated")))
                dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
            except Exception:
                dt = None
            if cutoff_date is not None and dt is not None and dt < cutoff_date:
                continue
            rating_raw = _label(e.get("im:rating"))
            author_field = e.get("author")
            author = _label(author_field.get("name")) if isinstance(author_field, dict) else _label(author_field)
            results.append({
                "id": rid,
                "date": dt,
                "rating": int(rating_raw) if rating_raw else None,
                "title": _label(e.get("title")),
                "text": _label(e.get("content")),
                "author": author,
                "version": _label(e.get("im:version")),
                "country": country,
            })
        except Exception as exc:
            logging.warning("App Store %s: пропущен нечитаемый отзыв: %s", country, exc)
    return results


def fetch_google_play_reviews(country, seen_hashes, limit=REVIEW_LIMIT, cutoff_date=None):
    """Без `cutoff_date`: забирает последние `limit` отзывов (обычный
    ежедневный режим, один запрос). С `cutoff_date`: постранично листает
    (сортировка по новизне), пока не выйдет за дату или не кончится
    история — режим для разовой догрузки истории на глубину N дней."""
    results = []
    try:
        if cutoff_date is None:
            batches = [gp_reviews(GOOGLE_PLAY_ID, lang="ru", country=country, sort=Sort.NEWEST, count=limit)[0]]
        else:
            batches = []
            continuation_token = None
            while True:
                batch, continuation_token = gp_reviews(
                    GOOGLE_PLAY_ID, lang="ru", country=country,
                    sort=Sort.NEWEST, count=200, continuation_token=continuation_token,
                )
                if not batch:
                    break
                batches.append(batch)
                oldest_at = batch[-1].get("at")
                if continuation_token is None or (oldest_at and oldest_at < cutoff_date):
                    break
                time.sleep(0.1)

        for batch in batches:
            for r in batch:
                author = r.get("userName", "") or ""
                text = r.get("content", "") or ""
                at = r.get("at")
                if cutoff_date is not None and at is not None and at < cutoff_date:
                    continue
                h = hash_review(author, at, text)
                if h in seen_hashes:
                    continue
                results.append({
                    "id": h,
                    "date": at,
                    "rating": r.get("score"),
                    "title": "",
                    "text": text,
                    "author": author,
                    "version": r.get("reviewCreatedVersion", "") or "",
                    "country": country,
                })
    except Exception as e:
        logging.warning("Google Play %s: %s", country, e)
    return results


PLATFORM_HEADER_FILL = PatternFill(start_color="C6D9F1", end_color="C6D9F1", fill_type="solid")
PLATFORM_HEADER_FONT = Font(bold=True, size=13)

DATE_IDX_F = FIELDS.index("Дата отзыва")
TITLE_IDX = FIELDS.index("Заголовок отзыва")
TEXT_IDX = FIELDS.index("Текст отзыва")
AUTHOR_IDX = FIELDS.index("Автор")

DATE_COLUMN_WIDTH = 22
TEXT_COLUMN_WIDTH = 40
AUTHOR_COLUMN_WIDTH = 15

# Колонка с заголовком отзыва только на стороне Apple (N) — с переносом.
WRAP_TITLE_COL = RIGHT_COL + TITLE_IDX

TEXT_ALIGNMENT = Alignment(wrap_text=True, vertical="top")  # E, O — как есть, без центрирования
AUTHOR_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)  # F, P
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CENTER_WRAP_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Обратная совместимость: старое имя, использовавшееся для E/O.
WRAP_ALIGNMENT = TEXT_ALIGNMENT


def data_cell_alignment(col):
    """Возвращает нужное выравнивание для колонки данных по её роли."""
    if col == LEFT_COL + TEXT_IDX or col == RIGHT_COL + TEXT_IDX:
        return TEXT_ALIGNMENT
    if col == LEFT_COL + AUTHOR_IDX or col == RIGHT_COL + AUTHOR_IDX:
        return AUTHOR_ALIGNMENT
    if col == WRAP_TITLE_COL:
        return CENTER_WRAP_ALIGNMENT
    return CENTER_ALIGNMENT


def set_column_widths(ws):
    for base in (LEFT_COL, RIGHT_COL):
        ws.column_dimensions[get_column_letter(base + DATE_IDX_F)].width = DATE_COLUMN_WIDTH
        ws.column_dimensions[get_column_letter(base + TEXT_IDX)].width = TEXT_COLUMN_WIDTH
        ws.column_dimensions[get_column_letter(base + AUTHOR_IDX)].width = AUTHOR_COLUMN_WIDTH


def ensure_header_merges(ws):
    """Чинит объединение ячеек в строке 1 (баннер платформы), если его
    когда-то случайно сняли — идемпотентно, не трогает остальное."""
    existing = {str(r) for r in ws.merged_cells.ranges}
    left_range = f"{get_column_letter(LEFT_COL)}1:{get_column_letter(LEFT_COL + len(FIELDS) - 1)}1"
    right_range = f"{get_column_letter(RIGHT_COL)}1:{get_column_letter(RIGHT_COL + len(FIELDS) - 1)}1"
    if left_range not in existing:
        ws.merge_cells(left_range)
    if right_range not in existing:
        ws.merge_cells(right_range)


def style_header_cells(ws):
    """Обновляет шрифт/заливку/выравнивание шапки (строки 1-2), не трогая
    объединение ячеек. Вызывается на каждом запуске, чтобы формат шапки
    не «съезжал», даже если сама шапка была создана более ранней версией
    скрипта."""
    for col in (LEFT_COL, RIGHT_COL):
        cell = ws.cell(row=1, column=col)
        cell.font = PLATFORM_HEADER_FONT
        cell.fill = PLATFORM_HEADER_FILL

    for j in range(len(FIELDS)):
        for base in (LEFT_COL, RIGHT_COL):
            cell = ws.cell(row=2, column=base + j)
            cell.font = Font(bold=True)
            cell.alignment = HEADER_ALIGNMENT

    set_column_widths(ws)


def setup_headers(ws):
    """Создаёт строки 1-2 (значения + объединение ячеек) для нового файла
    или при миграции со старого формата листа."""
    ws.append([None] * TOTAL_COLS)
    ws.cell(row=1, column=LEFT_COL, value="Google Play (Android)")
    ws.merge_cells(start_row=1, start_column=LEFT_COL, end_row=1, end_column=LEFT_COL + len(FIELDS) - 1)
    ws.cell(row=1, column=RIGHT_COL, value="App Store (iOS)")
    ws.merge_cells(start_row=1, start_column=RIGHT_COL, end_row=1, end_column=RIGHT_COL + len(FIELDS) - 1)

    ws.append([None] * TOTAL_COLS)
    for j, name in enumerate(FIELDS):
        ws.cell(row=2, column=LEFT_COL + j, value=name)
        ws.cell(row=2, column=RIGHT_COL + j, value=name)

    style_header_cells(ws)


def load_or_create_workbook():
    if OUTPUT_FILE.exists():
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Отзывы"
        setup_headers(ws)
    return wb, ws


def is_new_layout(ws):
    return ws.cell(row=1, column=LEFT_COL).value == "Google Play (Android)"


def load_existing(ws):
    """Возвращает (android_rows, apple_rows) в новом 9-польном формате.
    Понимает и старый плоский формат листа — для миграции ранее
    накопленных данных при первом запуске с новым кодом."""
    android_rows, apple_rows = [], []

    if is_new_layout(ws):
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row:
                continue
            left = list(row[LEFT_COL - 1:LEFT_COL - 1 + len(FIELDS)])
            right = list(row[RIGHT_COL - 1:RIGHT_COL - 1 + len(FIELDS)])
            if left[ID_IDX]:
                android_rows.append(left)
            if right[ID_IDX]:
                apple_rows.append(right)
    else:
        store_idx = OLD_HEADERS.index("Стор")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= store_idx:
                continue
            store = row[store_idx]
            if store not in ("App Store", "Google Play"):
                continue
            # OLD_HEADERS без "Стор" (индекс 1) даёт ровно порядок FIELDS.
            new_row = [row[0]] + list(row[2:])
            (android_rows if store == "Google Play" else apple_rows).append(new_row)

    return android_rows, apple_rows


def build_row(item):
    rating = item.get("rating")
    is_negative = rating is not None and rating <= 3
    row = [
        item.get("date"), item.get("country"), rating,
        item.get("title", ""), item.get("text", ""), item.get("author", ""),
        item.get("version", ""), item["id"], "да" if is_negative else "нет",
    ]
    return row, is_negative


DATE_BLOCK_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
DATE_BLOCK_FONT = Font(bold=True, size=12)
SECTION_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
SECTION_FONT = Font(bold=True, size=13)
ARCHIVE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

def _merged_banner_row(ws, text, fill, font):
    ws.append([None] * TOTAL_COLS)
    r = ws.max_row
    ws.cell(row=r, column=1, value=text)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS)
    cell = ws.cell(row=r, column=1)
    cell.fill = fill
    cell.font = font
    return r


def _group_by_day(rows):
    rows = sorted(rows, key=lambda r: r[DATE_IDX] or datetime.min, reverse=True)
    by_day = {}
    for day, day_rows in groupby(rows, key=lambda r: r[DATE_IDX].date() if hasattr(r[DATE_IDX], "date") else None):
        by_day[day] = list(day_rows)
    return by_day


def _write_day_block(ws, day, android_day_rows, apple_day_rows):
    label = day.strftime("%d.%m.%Y") if day else "Без даты"
    _merged_banner_row(ws, label, DATE_BLOCK_FILL, DATE_BLOCK_FONT)

    n = max(len(android_day_rows), len(apple_day_rows))
    for i in range(n):
        row_values = [None] * TOTAL_COLS
        a_row = android_day_rows[i] if i < len(android_day_rows) else None
        i_row = apple_day_rows[i] if i < len(apple_day_rows) else None
        if a_row is not None:
            row_values[LEFT_COL - 1:LEFT_COL - 1 + len(FIELDS)] = a_row
        if i_row is not None:
            row_values[RIGHT_COL - 1:RIGHT_COL - 1 + len(FIELDS)] = i_row
        ws.append(row_values)
        r = ws.max_row
        for col in range(1, TOTAL_COLS + 1):
            if col == GAP_COL:
                continue
            ws.cell(row=r, column=col).alignment = data_cell_alignment(col)
        if a_row is not None and a_row[NEG_IDX] == "да":
            for col in range(LEFT_COL, LEFT_COL + len(FIELDS)):
                ws.cell(row=r, column=col).fill = NEGATIVE_FILL
        if i_row is not None and i_row[NEG_IDX] == "да":
            for col in range(RIGHT_COL, RIGHT_COL + len(FIELDS)):
                ws.cell(row=r, column=col).fill = NEGATIVE_FILL


def _write_section(ws, android_rows, apple_rows):
    android_by_day = _group_by_day(android_rows)
    apple_by_day = _group_by_day(apple_rows)
    all_days = sorted(
        set(android_by_day) | set(apple_by_day),
        key=lambda d: d or datetime.min.date(),
        reverse=True,
    )
    for day in all_days:
        _write_day_block(ws, day, android_by_day.get(day, []), apple_by_day.get(day, []))


def write_side_by_side(ws, existing_android, existing_apple, new_android, new_apple):
    """Полностью перезаписывает данные листа (кроме шапки в строках 1-2):
    Android слева, Apple справа, строки внутри дня выровнены по индексу,
    при расхождении в количестве — пустые ячейки. Сверху — секция «НОВЫЕ»
    (добавлено в этом запуске, ещё не рассылалось), ниже — «АРХИВ» (уже
    было в файле и уже отправлялось раньше)."""
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= 3:  # не трогаем объединения в шапке (строки 1-2)
            ws.unmerge_cells(str(merged_range))
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)
    ensure_header_merges(ws)
    style_header_cells(ws)

    new_count = len(new_android) + len(new_apple)
    if new_count:
        _merged_banner_row(ws, f"НОВЫЕ С ПОСЛЕДНЕЙ ОТПРАВКИ ({new_count})", SECTION_FILL, SECTION_FONT)
        _write_section(ws, new_android, new_apple)

    if existing_android or existing_apple:
        _merged_banner_row(ws, "АРХИВ (уже отправлялось ранее)", ARCHIVE_FILL, SECTION_FONT)
        _write_section(ws, existing_android, existing_apple)


def notify_negative(count):
    try:
        from winotify import Notification
        toast = Notification(
            app_id="Отзывы Halyk",
            title="Новые негативные отзывы",
            msg=f"Найдено {count} новых отзывов с рейтингом ≤3★",
            duration="long",
        )
        toast.show()
    except Exception:
        logging.exception("Не удалось показать toast-уведомление")


def main():
    import argparse
    from datetime import timedelta

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--days", type=int, default=None,
        help="Разовая догрузка истории на глубину N дней от текущей даты "
             "(вместо обычного топ-%d за запуск). Дополняет текущий файл, "
             "ничего не удаляет." % REVIEW_LIMIT,
    )
    args = parser.parse_args()
    lookback_days = args.days if args.days else DAILY_LOOKBACK_DAYS
    cutoff_date = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=lookback_days)

    wb, ws = load_or_create_workbook()
    # То, что уже было в файле до этого запуска, уже ушло в прошлой рассылке.
    existing_android, existing_apple = load_existing(ws)
    seen_gp_hashes = {str(r[ID_IDX]) for r in existing_android}
    seen_app_ids = {str(r[ID_IDX]) for r in existing_apple}

    new_android, new_apple = [], []
    new_total = 0
    new_negative = 0

    for country in COUNTRIES:
        try:
            for item in fetch_app_store_reviews(country, seen_app_ids, cutoff_date=cutoff_date):
                seen_app_ids.add(item["id"])
                new_total += 1
                row, is_negative = build_row(item)
                new_apple.append(row)
                if is_negative:
                    new_negative += 1
        except Exception:
            logging.exception("App Store %s: сбор не удался", country)

    for country in COUNTRIES:
        try:
            for item in fetch_google_play_reviews(country, seen_gp_hashes, cutoff_date=cutoff_date):
                seen_gp_hashes.add(item["id"])
                new_total += 1
                row, is_negative = build_row(item)
                new_android.append(row)
                if is_negative:
                    new_negative += 1
        except Exception:
            logging.exception("Google Play %s: сбор не удался", country)

    retention_cutoff = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=RETENTION_DAYS)
    before = len(existing_android) + len(existing_apple) + len(new_android) + len(new_apple)
    existing_android = [r for r in existing_android if r[DATE_IDX] is None or r[DATE_IDX] >= retention_cutoff]
    existing_apple = [r for r in existing_apple if r[DATE_IDX] is None or r[DATE_IDX] >= retention_cutoff]
    new_android = [r for r in new_android if r[DATE_IDX] is None or r[DATE_IDX] >= retention_cutoff]
    new_apple = [r for r in new_apple if r[DATE_IDX] is None or r[DATE_IDX] >= retention_cutoff]
    removed = before - len(existing_android) - len(existing_apple) - len(new_android) - len(new_apple)

    if not is_new_layout(ws):
        for merged_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged_range))
        ws.delete_rows(1, ws.max_row)
        setup_headers(ws)

    write_side_by_side(ws, existing_android, existing_apple, new_android, new_apple)
    wb.save(OUTPUT_FILE)
    logging.info(
        "Готово. Новых отзывов: %s, из них негативных: %s. Удалено старше %s дней: %s",
        new_total, new_negative, RETENTION_DAYS, removed,
    )

    if new_negative > 0:
        notify_negative(new_negative)

    # Публикует итоги для шага отправки письма в GitHub Actions (см. workflow).
    gh_output = os.environ.get("GITHUB_OUTPUT")
    if gh_output:
        with open(gh_output, "a", encoding="utf-8") as f:
            f.write(f"new_total={new_total}\nnew_negative={new_negative}\n")


if __name__ == "__main__":
    main()
